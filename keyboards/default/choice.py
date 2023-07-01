from aiogram.types import ReplyKeyboardMarkup, KeyboardButton
import openpyxl

wb = openpyxl.load_workbook('data/Excel_File.xlsx')
ws = wb.active
values = [ws.cell(row=i, column=1).value for i in range(2, ws.max_row + 1)]
maxrow = ws.max_row + 1 - 2
tel_type = "iphone"
for i in range(0, maxrow):
    if tel_type != values[i]:
        tel_type = values[i]

        phone_type = ReplyKeyboardMarkup(
            keyboard=[
                [
                    KeyboardButton(text="value[i]")
                ]
            ]

        )

phone_seria = ReplyKeyboardMarkup(
    keyboard=[
        [
            KeyboardButton(text="iPhone 14 Series"),
            KeyboardButton(text="iPhone 13 Series")
        ],
        [
            KeyboardButton(text="iPhone 12 Series"),
            KeyboardButton(text="iPhone 11 Series")
        ],

    ],
    resize_keyboard=True
)
sim_kart = ReplyKeyboardMarkup(
    keyboard=[
        [
            KeyboardButton(text='1 sim'),
            KeyboardButton(text='D sim'),
            KeyboardButton(text='E sim'),
        ]

    ],
    resize_keyboard=True
)

color = ReplyKeyboardMarkup(
    keyboard=[
        [
            KeyboardButton(text='Black'),
        ],
        [
            KeyboardButton(text='Purple'),
            KeyboardButton(text='Blue'),
        ],
        [
            KeyboardButton(text='Red'),
            KeyboardButton(text='White'),
        ],
        [
            KeyboardButton(text='Green'),
            KeyboardButton(text='Midnight')
        ]
    ], resize_keyboard=True
)

Memory = ReplyKeyboardMarkup(
    keyboard=[
        [
            KeyboardButton(text='64GB'),
            KeyboardButton(text='128GB'),
        ],
        [
            KeyboardButton(text='256GB'),
            KeyboardButton(text='512GB'),
            KeyboardButton(text='1GB'),
        ],
    ], resize_keyboard=True
)

bolib = ReplyKeyboardMarkup(
    keyboard=[
        [
            KeyboardButton(text="Bo'lib to'lash")
        ]
    ], resize_keyboard=True
)


oy_oy = ReplyKeyboardMarkup(
    keyboard=[
        [
            KeyboardButton(text="1 oy"),
            KeyboardButton(text="2 oy"),
            KeyboardButton(text="3 oy"),
            KeyboardButton(text="4 oy")
        ],
        [
            KeyboardButton(text="5 oy"),
            KeyboardButton(text="6 oy"),
            KeyboardButton(text="7 oy"),
            KeyboardButton(text="8 oy")
        ],
        [
            KeyboardButton(text="9 oy"),
            KeyboardButton(text="10 oy"),
            KeyboardButton(text="11 oy"),
            KeyboardButton(text="12 oy")
        ],

    ], resize_keyboard=True
)
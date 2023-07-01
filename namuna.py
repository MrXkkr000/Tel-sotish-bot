import openpyxl

wb = openpyxl.load_workbook('Price_list (2).xlsx')
ws = wb.active
# activ qiladi
# print('Total number of rows: '+str(ws.max_row)+'. And total number of columns: '+str(ws.max_column))
# # max cator va ustunni topadi
#
# print('The value in cell A1 is: '+ws['A1'].value)
# # A1 ni topadi

values = [ws.cell(row=i, column=1).value for i in range(2, ws.max_row+1)]
maxrow=ws.max_row+1-2
tel_type = "iphone"
for i in range(0, maxrow):
    if tel_type != values[i]:
        print(values[i])
        tel_type = values[i]
# print(values)
# ['iPhone 14', 'iPhone 14', 'iPhone 14', 'iPhone 14', 'iPhone 14', 'iPhone 14 Plus', 'iPhone 14 Plus', 'iPhone 14
# Plus' , 'iPhone 14 Plus', 'iPhone 14 Plus', 'iPhone 14 Plus', 'iPhone 14 Pro', 'iPhone 14 Pro', 'iPhone 14 Pro',
# 'iPhone 14 Pro', 'iPhone 14 Pro Max', 'iPhone 14 Pro Max', 'iPhone 14 Pro Max', 'iPhone 14 Pro Max', 'iPhone 13
# Mini', 'iPhone 13 Mini', 'iPhone 13 Mini', 'iPhone 13 Mini', 'iPhone 13 Mini', 'iPhone 13 Mini', 'iPhone 13',
# 'iPhone 13', 'iPhone 13', 'iPhone 13', 'iPhone 13', 'iPhone 13', 'iPhone 13 Pro', 'iPhone 13 Pro', 'iPhone 13 Pro',
# 'iPhone 13 Pro', 'iPhone 13 Pro', 'iPhone 13 Pro Max', 'iPhone 13 Pro Max', 'iPhone 13 Pro Max', 'iPhone 13 Pro
# Max', 'iPhone 13 Pro Max', 'iPhone 12 Mini', 'iPhone 12 Mini', 'iPhone 12 Mini', 'iPhone 12 Mini', 'iPhone 12
# Mini', 'iPhone 12 Mini', 'iPhone 12', 'iPhone 12', 'iPhone 12', 'iPhone 12', 'iPhone 12', 'iPhone 12', 'iPhone 12
# Pro', 'iPhone 12 Pro', 'iPhone 12 Pro', 'iPhone 12 Pro', 'iPhone 12 Pro Max', 'iPhone 12 Pro Max', 'iPhone 12 Pro
# Max', 'iPhone 12 Pro Max', 'iPhone 11', 'iPhone 11', 'iPhone 11', 'iPhone 11', 'iPhone 11']

# Bu tel types



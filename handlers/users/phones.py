import openpyxl
from aiogram import types
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters.state import StatesGroup, State
from aiogram.types import ReplyKeyboardRemove, InputFile

from keyboards.default.choice import phone_seria, bolib, oy_oy
from keyboards.default.hotira import gb1, gb64, gb256, gb512
from keyboards.default.phone_type import iPhone14, iPhone11, iPhone12, iPhone13
from loader import dp

wb = openpyxl.load_workbook('data/Excel_File.xlsx')
ws = wb.active
values_number = [ws.cell(row=i, column=1).value for i in range(2, ws.max_row + 1)]
values_series = [ws.cell(row=i, column=2).value for i in range(2, ws.max_row + 1)]
values_phone = [ws.cell(row=i, column=3).value for i in range(2, ws.max_row + 1)]
values_memory = [ws.cell(row=i, column=4).value for i in range(2, ws.max_row + 1)]
values_color = [ws.cell(row=i, column=6).value for i in range(2, ws.max_row + 1)]
values_sim = [ws.cell(row=i, column=7).value for i in range(2, ws.max_row + 1)]
values_narx = [ws.cell(row=i, column=8).value for i in range(2, ws.max_row + 1)]
maxrow = ws.max_row + 1 - 2
tel_type = "iphone"


# for i in range(2, maxrow):
#     if tel_type != values[i]:
#         tel_type = values[i]


class ClientData(StatesGroup):
    series = State()
    model = State()
    memory = State()
    color = State()
    sim = State()
    bolib = State()
    first_pay = State()
    oy = State()


@dp.message_handler(text='/start', state=None)
async def series_answer(message: types.Message, state: FSMContext):
    photo = InputFile('data/img.png')
    await message.answer_photo(photo, reply_markup=phone_seria)
    await ClientData.series.set()



@dp.message_handler(
    lambda message: message.text != "iPhone 14 Series" and message.text != "iPhone 13 Series" and message.text != "iPhone 12 Series" and message.text != "iPhone 11 Series",
    state=ClientData.series)
async def process_age_ivalid(message: types.Message):
    """
    If age is invalid
    """
    return await message.reply("To'g'ri bo'limni tanlang:")


@dp.message_handler(
    lambda message: message.text == "iPhone 14 Series" or message.text == "iPhone 13 Series" or message.text == "iPhone 12 Series" or message.text == "iPhone 11 Series",
    state=ClientData.series)
async def bot_tart(message: types.Message, state: FSMContext):
    seria = message.text

    await state.update_data(
        {"seria": seria}
    )
    # tel_type = "iphone"
    # for i in range(0, maxrow):
    #     if tel_type != values_phone[i] and seria in values_series[i]:
    #         tel_type = values_phone[i]
    #
    #         await message.answer(f"{i}.📱 {values_phone[i]}\n")

    if seria == "iPhone 14 Series":
        await message.answer(f"Modelni tanlang", reply_markup=iPhone14)
    if seria == "iPhone 13 Series":
        await message.answer(f"Modelni tanlang", reply_markup=iPhone13)
    if seria == "iPhone 12 Series":
        await message.answer(f"Modelni tanlang", reply_markup=iPhone12)
    if seria == "iPhone 11 Series":
        await message.answer(f"Modelni tanlang", reply_markup=iPhone11)
    await ClientData.next()


@dp.message_handler(
    lambda message: message.text != "iPhone 11" and message.text != "iPhone 12" and message.text != "iPhone 12 Mini" and message.text != "iPhone 13 Mini"
                    and message.text != "iPhone 12 Pro Max" and message.text != "iPhone 12 Pro" and message.text != "iPhone 13" and message.text != "iPhone 14 Plus"
                    and message.text != "iPhone 13 Pro Max" and message.text != "iPhone 13 Pro" and message.text != "iPhone 14 Pro Max" and message.text != "iPhone 14 Pro" and message.text != "iPhone 14",
    state=ClientData.model)
async def process_ageid(message: types.Message):
    """
    If age is invalid
    """
    return await message.reply("To'g'ri bo'limni tanlang:")


@dp.message_handler(state=ClientData.model)
async def answer_model(message: types.Message, state: FSMContext):
    model = message.text
    await state.update_data(
        {"model": model}
    )
    if model == "iPhone 11":
        await message.answer(f"Xotira hajmini tanlang", reply_markup=gb64)

    elif model == "iPhone 12" or model == "iPhone 12 Mini" or model == "iPhone 13 Mini":
        await message.answer(f"Xotira hajmini tanlang", reply_markup=gb256)

    elif model == "iPhone 12 Pro Max" or model == "iPhone 12 Pro" or model == "iPhone 13" or model == "iPhone 14 " \
                                                                                                      "Plus" or \
            model == "iPhone 14":
        await message.answer(f"Xotira hajmini tanlang", reply_markup=gb512)

    elif model == "iPhone 13 Pro Max" or model == "iPhone 13 Pro" or model == "iPhone 14 Pro Max" or model == "iPhone 14 Pro":
        await message.answer(f"Xotira hajmini tanlang", reply_markup=gb1)

    await ClientData.next()


@dp.message_handler(
    lambda message: message.text != "64 GB" and message.text != "128 GB" and message.text != "256 GB" and message.text != "512 GB" and message.text != "1 TB",
    state=ClientData.memory)
async def process_age_ild(message: types.Message):
    """
    If age is invalid
    """
    return await message.reply("Bu yerga kerakli bo'limni tanlang:")


@dp.message_handler(
    lambda message: message.text != "64 GB" or message.text == "128 GB" or message.text == "256 GB" or message.text == "512 GB" or message.text == "1 TB",
    state=ClientData.memory)
async def awr_mode(message: types.Message, state: FSMContext):
    memory = message.text
    await state.update_data(
        {"memory": memory}
    )
    data = await state.get_data()
    seria = data.get("seria")
    model = data.get("model")
    memory = data.get("memory")
    tel_color = "oq"
    await message.answer(f"Telefon rangi raqamini kiriting")
    for i in range(0, maxrow):
        if seria == values_series[i] and tel_color != values_color[i] and model == values_phone[i] and memory == \
                values_memory[i]:
            # and model == values_phone[i] and memory == values_memory[i] and tel_color != values_color[i]:
            tel_color = values_color[i]
            await message.answer(f"{i}. {values_color[i]}\n", reply_markup=ReplyKeyboardRemove())
    await ClientData.next()


@dp.message_handler(lambda message: not message.text.isdigit(), state=ClientData.color)
async def process_age_ild(message: types.Message):
    """
    If age is invalid
    """
    return await message.reply("Bu yerga faqat raqam kiriting( rang raqami ):\nMassalan: 125")


@dp.message_handler(lambda message: message.text.isdigit(), state=ClientData.color)
async def answer_moel(message: types.Message, state: FSMContext):
    nomer = int(message.text)
    color = values_color[nomer]
    await state.update_data(
        {"color": color}
    )
    data = await state.get_data()
    seria = data.get("seria")
    model = data.get("model")
    memory = data.get("memory")
    color = data.get("color")
    tel_sim = "sim"
    await message.answer(f"Sim karta turi raqamini kiriting")
    for i in range(0, maxrow):
        if seria == values_series[i] and model == values_phone[i] and memory == values_memory[i] and color == \
                values_color[i] and tel_sim != values_sim[i]:
            # and model == values_phone[i] and memory == values_memory[i] and tel_color != values_color[i]:
            tel_sim = values_sim[i]
            await message.answer(f"{i}. {values_sim[i]}\n")
    await ClientData.next()


@dp.message_handler(lambda message: not message.text.isdigit(), state=ClientData.sim)
async def proc_invalid(message: types.Message):
    """
    If age is invalid
    """
    return await message.reply("Bu yerga faqat raqam kiriting( sim raqami ):\nMassalan: 125")


@dp.message_handler(lambda message: message.text.isdigit(), state=ClientData.sim)
async def anmodel(message: types.Message, state: FSMContext):
    nomers = int(message.text)
    sim = values_sim[nomers]
    await state.update_data(
        {"sim": sim}
    )
    data = await state.get_data()
    seria = data.get("seria")
    model = data.get("model")
    memory = data.get("memory")
    color = data.get("color")
    sim = data.get("sim")
    for i in range(0, maxrow):
        if seria == values_series[i] and model == values_phone[i] and (memory) == (values_memory[i]) and color == \
                values_color[i] and sim == values_sim[i]:
            narx = values_narx[i]
            await message.answer(
                f"📱Model                       : {model}\n🗂Xotira                       : {memory}\n🖌Rang                         : {color}\n🎫Sim karta turi       : {sim}\n🧮Umumiy narx          : {narx} $",
                reply_markup=bolib)
    await ClientData.next()


@dp.message_handler(state=ClientData.bolib)
async def asw_ml(message: types.Message, state: FSMContext):
    await message.answer(f"Boshlang'ich to'lovni kiriting ($ Dollarda)",
                         reply_markup=ReplyKeyboardRemove())
    await ClientData.next()


@dp.message_handler(lambda message: not message.text.isdigit(), state=ClientData.first_pay)
async def ps_age_invalid(message: types.Message):
    """
    If age is invalid
    """
    return await message.reply("Bu yerga faqat raqam kiriting( $ dollorda ):\nMassalan: 150")


@dp.message_handler(lambda message: message.text.isdigit(), state=ClientData.first_pay)
async def answdel(message: types.Message, state: FSMContext):
    f_p = int(message.text)
    await state.update_data(
        {"f_p": f_p}
    )
    await message.answer(f"To'lov muddatini tanlang", reply_markup=oy_oy)
    await ClientData.next()


# @dp.message_handler(lambda message: not message.text.isdigit(), state=ClientData.oy)
# async def process_age_invalid(message: types.Message):
#     """
#     If age is invalid
#     """
#     return await message.reply("Bu yerga faqat raqam kiriting( oy ):\nMassalan: 12")

@dp.message_handler(
    lambda message: message.text != "1 oy" and message.text != "2 oy" and message.text != "3 oy" and message.text != "4 oy" and message.text != "5 oy" and message.text != "6 oy" and message.text != "7 oy" and message.text != "8 oy" and message.text != "9 oy" and message.text != "10 oy" and message.text != "11 oy" and message.text != "12 oy",
    state=ClientData.oy)
async def process_age_ild(message: types.Message):
    """
    If age is invalid
    """
    return await message.reply("Bu yerga kerakli bo'limni tanlang:")


@dp.message_handler(
    lambda message: message.text == "1 oy" or message.text == "2 oy" or message.text == "3 oy" or message.text == "4 oy" or message.text == "5 oy" or message.text == "6 oy" or message.text == "7 oy" or message.text == "8 oy" or message.text == "9 oy" or message.text == "10 oy" or message.text == "11 oy" or message.text == "12 oy",
    state=ClientData.oy)
async def aoodel(message: types.Message, state: FSMContext):

    oy_1 = message.text
    if oy_1 == "1 oy":
        oy = 1
    elif oy_1 == "2 oy":
        oy = 2
    elif oy_1 == "3 oy":
        oy = 3
    elif oy_1 == "4 oy":
        oy = 4
    elif oy_1 == "5 oy":
        oy = 5
    elif oy_1 == "6 oy":
        oy = 6
    elif oy_1 == "7 oy":
        oy = 7
    elif oy_1 == "8 oy":
        oy = 8
    elif oy_1 == "9 oy":
        oy = 9
    elif oy_1 == "10 oy":
        oy = 10
    elif oy_1 == "11 oy":
        oy = 11
    elif oy_1 == "12 oy":
        oy = 12

    await state.update_data(
        {"oy": oy}
    )
    data = await state.get_data()
    seria = data.get("seria")
    model = data.get("model")
    memory = data.get("memory")
    color = data.get("color")
    sim = data.get("sim")
    f_p = int(data.get("f_p"))
    oy = int(data.get("oy"))

    await message.answer(f"Sizning malumotlaringiz", reply_markup=ReplyKeyboardRemove())
    for i in range(0, maxrow):
        if seria == values_series[i] and model == values_phone[i] and memory == (values_memory[i]) and color == \
                values_color[i] and sim == values_sim[i]:
            narx = int(values_narx[i])
            ustama = 0
            foiz_b_t = int((f_p * 100) / narx)
            if foiz_b_t >= 50:
                if oy == 3 or oy == 4:
                    ustama = 5
                elif oy == 5 or oy == 6:
                    ustama = 6
                elif oy == 7 or oy == 8:
                    ustama = 7
                elif oy >= 9 and oy <= 12:
                    ustama = 8


            else:
                if oy == 3 or oy == 4:
                    ustama = 6
                elif oy == 5 or oy == 6:
                    ustama = 7
                elif oy == 7 or oy == 8:
                    ustama = 8
                elif oy >= 9 and oy <= 12:
                    ustama = 9
            narx_qo = narx-f_p
            narx_qo += narx_qo * oy*(ustama / 100)
            ustama = 0
            oyiga = narx_qo/oy


            await message.answer(
                f"📱Model                       : {model}\n🗂Xotira                       : {memory}\n🖌Rang                         : {color}\n🎫Sim karta turi           : {sim}\n🧮Umumiy narx          : {narx} $\n🎯Boshlang'ich to'lov: {f_p} $ \n🔄Muddat                        : {oy} oy \n📉To'lov oyiga             : {round(oyiga, 2)} $ ")
    await state.finish()
